import os
import sys
import glob

# Ensure dependencies are installed
try:
    import pandas
    import openpyxl
except ImportError:
    import subprocess
    print("Installing missing dependencies...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl", "--break-system-packages"])

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Directory containing the CSVs
dir_path = "/Users/Luthfi/Project/PT. Teknalogi Transformasi Digital/PertaminaWebCareer/Folder Pertamina New"
csv_files = glob.glob(os.path.join(dir_path, "loker_magang_pertamina_semua*.csv"))

def get_page_number(x):
    name = os.path.basename(x)
    name_no_ext = name.split('.')[0]
    parts = name_no_ext.split('_')
    if parts[-1].isdigit():
        return int(parts[-1])
    return 1 # Fallback for loker_magang_pertamina_semua.csv (page 1)

# Sort files by page number
csv_files.sort(key=get_page_number)

dfs = []
for file in csv_files:
    try:
        df = pd.read_csv(file, encoding='utf-8-sig')
        # Add a column for source page to help in auditing
        df['Source_Page'] = get_page_number(file)
        dfs.append(df)
    except Exception as e:
        print(f"Error reading {file}: {e}")

if not dfs:
    print("No CSV files found!")
    sys.exit(1)

combined_df = pd.concat(dfs, ignore_index=True)

# ----------------- AUDIT PROCESS -----------------
total_raw = len(combined_df)

# Check duplicates by 'Link Detail'
duplicate_mask = combined_df.duplicated(subset=['Link Detail'], keep=False)
duplicates_df = combined_df[duplicate_mask]

# Remove duplicates, keeping the first occurrence
dedup_df = combined_df.drop_duplicates(subset=['Link Detail'], keep='first')
total_unique = len(dedup_df)
total_duplicates_removed = total_raw - total_unique

# Create Audit Report text
report = []
report.append("====================================================")
report.append("            AUDIT REPORT - LOKER PERTAMINA 1-40")
report.append("====================================================")
report.append(f"Total Lowongan Sebelum Deduplikasi : {total_raw}")
report.append(f"Total Lowongan Unik                : {total_unique}")
report.append(f"Total Duplikat yang Dihapus        : {total_duplicates_removed}")
report.append("====================================================")

# Breakdown by Perusahaan
report.append("\nBreakdown Lowongan Berdasarkan Perusahaan (Top 10):")
perusahaan_counts = dedup_df['Perusahaan'].value_counts()
for name, val in perusahaan_counts.head(10).items():
    report.append(f" - {str(name):<45} : {val} lowongan")

# Breakdown by Kota
report.append("\nBreakdown Lowongan Berdasarkan Kota (Top 10):")
kota_counts = dedup_df['Kota'].value_counts()
for name, val in kota_counts.head(10).items():
    report.append(f" - {str(name):<45} : {val} lowongan")

# Breakdown by Pendidikan
report.append("\nBreakdown Lowongan Berdasarkan Pendidikan:")
pendidikan_counts = dedup_df['Pendidikan'].value_counts()
for name, val in pendidikan_counts.items():
    report.append(f" - {str(name):<45} : {val} lowongan")

# List duplicates detail if any
if total_duplicates_removed > 0:
    report.append("\nDetail Duplikasi yang Ditemukan (Beberapa Contoh):")
    # Show first few duplicates
    examples = duplicates_df.sort_values(by='Link Detail').head(10)
    for idx, row in examples.iterrows():
        report.append(f" - Halaman {row['Source_Page']}: {row['Judul Lowongan']} ({row['Perusahaan']})")

# Write report to text file
report_text = "\n".join(report)
print(report_text)

with open(os.path.join(dir_path, "audit_summary.txt"), "w", encoding="utf-8") as f:
    f.write(report_text)

# ----------------- CREATE EXCEL -----------------
# Remove Source_Page from final output to keep it clean
output_df = dedup_df.drop(columns=['Source_Page']).copy()
output_df.sort_values(by=['Perusahaan', 'Judul Lowongan'], inplace=True)

excel_path = os.path.join(dir_path, "loker_magang_pertamina_page_1-40.xlsx")

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    output_df.to_excel(writer, sheet_name="Loker Magang 1-40", index=False)
    
    workbook = writer.book
    sheet = writer.sheets["Loker Magang 1-40"]
    sheet.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    
    # Header: Pertamina Blue theme
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(name=font_family, size=10)
    data_align = Alignment(vertical="center", wrap_text=False)
    data_align_wrap = Alignment(vertical="center", wrap_text=True)
    
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Header Styling
    for col_idx, col in enumerate(output_df.columns, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    # Data rows styling
    for row_idx in range(2, sheet.max_row + 1):
        is_even = (row_idx % 2 == 0)
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
            # Wrap text for columns 1 (Judul), 7 (Jurusan)
            if col_idx in [1, 7]:
                cell.alignment = data_align_wrap
            else:
                cell.alignment = data_align
                
            if is_even:
                cell.fill = zebra_fill
                
    # Auto adjust column widths
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith("http"):
                max_len = max(max_len, 15)
            else:
                lines = val_str.split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
        sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

print(f"\nSaved beautiful Excel sheet to: {excel_path}")
